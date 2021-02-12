import face_recognition
from PIL import Image, ImageDraw
from datetime import datetime
from openpyxl import load_workbook

image_of_bill = face_recognition.load_image_file('./img/known/Bill Gates.jpg')
bill_face_encoding = face_recognition.face_encodings(image_of_bill)[0]

image_of_steve = face_recognition.load_image_file('./img/known/Steve Jobs.jpg')
steve_face_encoding = face_recognition.face_encodings(image_of_steve)[0]

image_of_elon = face_recognition.load_image_file('./img/known/Elon Musk.jpg')
elon_face_encoding = face_recognition.face_encodings(image_of_elon)[0]

#  Create arrays of encodings and names
known_face_encodings = [
  bill_face_encoding,
  steve_face_encoding,
  elon_face_encoding
]

known_face_names = [
  "Bill Gates",
  "Steve Jobs",
  "Elon Musk"
]

# Load test image to find faces in
test_image = face_recognition.load_image_file('./img/known/Elon Musk.jpg')

# Find faces in test image
face_locations = face_recognition.face_locations(test_image)
face_encodings = face_recognition.face_encodings(test_image, face_locations)


# Setting DB req ready
workbook = load_workbook(filename="Attendance.xlsx")
sheet = workbook["Class-1A"]
lastRow = sheet.max_row
Studentnames = [ i.value for i in sheet[1] ]
currentDate = datetime.today().strftime('%d/%m/%Y')


# Loop through faces in test image
for(top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):
  matches = face_recognition.compare_faces(known_face_encodings, face_encoding)

  name = "Unknown Person"

  # If match
  if True in matches:
    first_match_index = matches.index(True)
    name = known_face_names[first_match_index]
  
  # Update in DB  
  if( sheet.cell(lastRow,Studentnames.index(name)+1).value == None):
    print(type(sheet.cell(lastRow,1).value))
    if(sheet.cell(lastRow,1).value == None):
      sheet.cell(lastRow,1).value = currentDate
      sheet.cell(lastRow,Studentnames.index(name)+1).value = True
      print(sheet.cell(lastRow,Studentnames.index(name)+1).value)

    elif(sheet.cell(lastRow,1).value == currentDate):      
      sheet.cell(lastRow,Studentnames.index(name)+1).value = True
      print(sheet.cell(lastRow,Studentnames.index(name)+1).value)

    elif(sheet.cell(lastRow,1).value != currentDate):
      sheet.cell(lastRow+1,1).value = currentDate
      sheet.cell(lastRow+1,Studentnames.index(name)+1).value = True
      print(sheet.cell(lastRow+1,Studentnames.index(name)+1).value)

  

print(lastRow)
workbook.save(filename="Attendance.xlsx")