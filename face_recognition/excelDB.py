from datetime import datetime
from openpyxl import load_workbook
workbook = load_workbook(filename="Attendance.xlsx")
#print(workbook.sheetnames) 

sheet = workbook["Class-1A"]
#print(sheet.title)
#print(sheet["A1"].value)

# Returns last row with data
#print(sheet.max_row)

lastRow = sheet.max_row
#print(sheet.cell(lastRow,1).value)

#gets today date
print(datetime.today().strftime('%d-%m-%Y'))

# Gets all data of row 
# names = [i.value for i in sheet[lastRow] ]
# print(names)

names = [i.value for i in sheet[1] ]

sheet.cell(lastRow+1,1).value = datetime.today().strftime('%d-%m-%Y')
indexVal = names.index("Bill Gates")
sheet.cell(lastRow+1,indexVal+1).value = True
workbook.save(filename="Attendance.xlsx")
